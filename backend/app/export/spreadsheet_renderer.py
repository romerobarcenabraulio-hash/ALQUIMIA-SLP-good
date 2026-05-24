"""
Fase 4 — spreadsheet_renderer.py

Renderiza modelos financieros y Gantt a XLSX institucional.

Reglas:
- Inputs separados de outputs.
- Fuentes visibles en hoja "Fuentes".
- Warnings visibles en hoja "Advertencias".
- Celdas sin dato → "N/D (ver simulador)" — nunca inventar cifras.
- Proteger contra hoja vacía: siempre headers + al menos fila de ejemplo.
- No hardcodear valores financieros si vienen del manifest/resultados.
"""
from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ─── Paleta de colores ALQUIMIA ───────────────────────────────────────────────

_GREEN  = "3B6D11"
_BLUE   = "1A5FA8"
_ORANGE = "D4881E"
_CREAM  = "F8F6F1"
_DARK   = "1C1B18"
_GREY   = "A8A49C"


def _header_fill(ws, color: str):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _bold_font(size: int = 11, color: str = "000000"):
    from openpyxl.styles import Font
    return Font(bold=True, size=size, color=color)


def _normal_font(size: int = 10, color: str = _DARK):
    from openpyxl.styles import Font
    return Font(bold=False, size=size, color=color)


def _set_header_row(ws, headers: list[str], color: str = _GREEN) -> None:
    from openpyxl.styles import Alignment
    fill = _header_fill(ws, color)
    font = _bold_font(color="FFFFFF")
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18


def _auto_width(ws, min_w: int = 12, max_w: int = 50) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(
            max(length + 2, min_w), max_w
        )


def _nd(val: Any) -> str:
    """Retorna el valor o 'N/D (ver simulador)' si es None/vacío."""
    if val is None or val == "" or val == {}:
        return "N/D (ver simulador)"
    return val


# ─── 05 Modelo Financiero CFO ────────────────────────────────────────────────

def render_financial_xlsx(
    manifest: dict,
    resultados: Optional[dict] = None,
    theme_zm: str = "",
    theme_municipio: str = "",
    package_id: str = "",
) -> bytes:
    """
    Genera 05_Modelo_Financiero_CFO.xlsx con hojas:
      Resumen, Inputs, Resultados, Flujo_Anual, Sensibilidades, Fuentes, Advertencias

    manifest: ExportManifest serializado como dict
    resultados: dict con KPIs del simulador (opcional — si falta, celdas N/D)
    """
    import openpyxl
    wb = openpyxl.Workbook()
    res = resultados or {}

    # ── Hoja: Resumen ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    _set_header_row(ws, ["Campo", "Valor"])
    summary_rows = [
        ("Zona Metropolitana",    manifest.get("zm") or theme_zm or "N/D"),
        ("Municipios",            ", ".join(manifest.get("municipios") or []) or "N/D"),
        ("Versión del paquete",   manifest.get("version", "0.1-borrador")),
        ("Fecha de exportación",  date.today().isoformat()),
        ("Score de datos",        f"{manifest.get('score_datos', 'N/D')}%" if manifest.get("score_datos") else "N/D"),
        ("Package ID",            package_id[:24] + "…" if len(package_id) > 24 else package_id),
        ("Fuentes utilizadas",    ", ".join(manifest.get("fuentes_usadas") or []) or "N/D"),
        ("Advertencias activas",  str(len(manifest.get("warnings_activos") or []))),
        ("Documentos KPI",        str(len(manifest.get("kpis_incluidos") or []))),
    ]
    for r, (k, v) in enumerate(summary_rows, 2):
        ws.cell(row=r, column=1, value=k).font = _bold_font(10, _DARK)
        ws.cell(row=r, column=2, value=v).font = _normal_font()
    _auto_width(ws)

    # ── Hoja: Inputs ──────────────────────────────────────────────────────────
    ws_in = wb.create_sheet("Inputs")
    _set_header_row(ws_in, ["KPI / Parámetro", "Valor ingresado", "Unidad", "Fuente / Nota"], _BLUE)
    input_fields = [
        ("RSU total ton/día",         res.get("rsu_total_ton_dia"),   "t/día",   "Simulador ALQUIMIA"),
        ("Horizonte de evaluación",   res.get("horizonte"),           "años",    "Configuración usuario"),
        ("WACC",                      res.get("wacc"),                "%",       "Configuración usuario"),
        ("Tipo de cambio USD/MXN",    res.get("tipo_cambio"),         "MXN/USD", "Fuente mercado"),
        ("Precio carbono escenario",  res.get("precio_carbono_esc"),  "scenario","Configuración usuario"),
        ("Merma logística %",         res.get("merma_log_pct"),       "%",       "Configuración usuario"),
        ("Costo comunicación social", res.get("costo_com_social"),    "MXN",     "Configuración usuario"),
    ]
    for r, (label, val, unit, nota) in enumerate(input_fields, 2):
        ws_in.cell(row=r, column=1, value=label)
        ws_in.cell(row=r, column=2, value=_nd(val))
        ws_in.cell(row=r, column=3, value=unit)
        ws_in.cell(row=r, column=4, value=nota)
    _auto_width(ws_in)

    # ── Hoja: Resultados ──────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resultados")
    _set_header_row(ws_res, ["Indicador", "Valor", "Unidad", "Nota"], _GREEN)
    result_fields = [
        ("TIR del proyecto",          res.get("tir"),                 "%",          "Retorno interno del proyecto"),
        ("TIR equity",                res.get("tir_equity"),          "%",          "Retorno para inversionista"),
        ("VPN",                       res.get("vpn"),                 "MXN",        "Valor presente neto"),
        ("CAPEX total",               res.get("capex_total"),         "MXN",        "Inversión inicial total"),
        ("EBITDA anual",              res.get("ebitda"),              "MXN/año",    "Resultado antes de depreciación"),
        ("Margen EBITDA",             res.get("margen_ebitda"),       "%",          ""),
        ("Payback (meses)",           res.get("payback_meses"),       "meses",      ""),
        ("MOIC",                      res.get("moic"),                "x",          "Múltiplo sobre inversión"),
        ("Ingresos brutos (horizonte)",res.get("ingresos_brutos"),    "MXN",        ""),
        ("Empleos directos",          res.get("empleos_directos"),    "personas",   ""),
        ("Empleos indirectos",        res.get("empleos_indirectos"),  "personas",   ""),
        ("CO2e evitadas/año",         res.get("co2e_evitadas_anual"), "t CO2e/año", "Año final del horizonte"),
        ("CO2e evitadas (horizonte)", res.get("co2e_evitadas_horizonte"), "t CO2e", "Acumulado"),
        ("kWh biogás",                res.get("kwh_biogas"),          "kWh/año",    ""),
        ("Extensión vida relleno",    res.get("extension_relleno"),   "años",       "Cap. 15 años"),
        ("Ahorro salud",              res.get("ahorro_salud"),        "MXN/año",    ""),
        ("Derrama total",             res.get("derrama_total"),       "MXN",        ""),
        ("Score político",            res.get("score_politico"),      "0-100",      ""),
    ]
    for r, (label, val, unit, nota) in enumerate(result_fields, 2):
        ws_res.cell(row=r, column=1, value=label)
        ws_res.cell(row=r, column=2, value=_nd(val))
        ws_res.cell(row=r, column=3, value=unit)
        ws_res.cell(row=r, column=4, value=nota)
    _auto_width(ws_res)

    # ── Hoja: Flujo_Anual ─────────────────────────────────────────────────────
    ws_fa = wb.create_sheet("Flujo_Anual")
    _set_header_row(ws_fa,
        ["Año", "% Captura", "Ingresos (MXN)", "CAPEX (MXN)", "OPEX (MXN)",
         "EBITDA (MXN)", "FCF (MXN)", "FCF acumulado (MXN)", "CO2e (t/año)"],
        _DARK
    )
    serie = res.get("serie_anual") or []
    if serie:
        for r, row_data in enumerate(serie, 2):
            ws_fa.cell(row=r, column=1, value=row_data.get("año", r - 1))
            ws_fa.cell(row=r, column=2, value=row_data.get("pctCaptura") or row_data.get("pct_captura"))
            ws_fa.cell(row=r, column=3, value=row_data.get("ingresos"))
            ws_fa.cell(row=r, column=4, value=row_data.get("capex"))
            ws_fa.cell(row=r, column=5, value=row_data.get("opex"))
            ws_fa.cell(row=r, column=6, value=row_data.get("ebitda"))
            ws_fa.cell(row=r, column=7, value=row_data.get("fcf"))
            ws_fa.cell(row=r, column=8, value=row_data.get("fcfAcumulado") or row_data.get("fcf_acumulado"))
            ws_fa.cell(row=r, column=9, value=row_data.get("co2e"))
    else:
        ws_fa.cell(row=2, column=1, value="N/D")
        ws_fa.cell(row=2, column=2, value="La serie anual no está disponible en este paquete.")
        ws_fa.cell(row=2, column=2).font = _normal_font(color=_GREY)
    _auto_width(ws_fa)

    # ── Hoja: Sensibilidades ──────────────────────────────────────────────────
    ws_sens = wb.create_sheet("Sensibilidades")
    _set_header_row(ws_sens, ["Variable", "Escenario base", "Escenario bajo", "Escenario alto", "Impacto en TIR"], _ORANGE)
    ws_sens.cell(row=2, column=1, value="Score de datos del paquete")
    ws_sens.cell(row=2, column=2, value=f"{manifest.get('score_datos', 'N/D')}%")
    ws_sens.cell(row=2, column=3, value="N/D (ver advertencias)")
    ws_sens.cell(row=2, column=4, value="N/D (ver advertencias)")
    ws_sens.cell(row=2, column=5, value="Depende del score de datos")
    ws_sens.cell(row=3, column=1, value="Nota")
    ws_sens.cell(row=3, column=2,
        value="Las sensibilidades requieren corrida de escenarios en el simulador ALQUIMIA.")
    ws_sens.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    _auto_width(ws_sens)

    # ── Hoja: Fuentes ─────────────────────────────────────────────────────────
    ws_fuentes = wb.create_sheet("Fuentes")
    _set_header_row(ws_fuentes, ["#", "Fuente", "Tipo", "Nota"], _BLUE)
    fuentes = manifest.get("fuentes_usadas") or ["Simulador ALQUIMIA"]
    for r, f in enumerate(fuentes, 2):
        ws_fuentes.cell(row=r, column=1, value=r - 1)
        ws_fuentes.cell(row=r, column=2, value=f)
        ws_fuentes.cell(row=r, column=3, value="Declarado en DataProvenance")
        ws_fuentes.cell(row=r, column=4, value="Ver manifest.json para detalle de confianza")
    _auto_width(ws_fuentes)

    # ── Hoja: Advertencias ────────────────────────────────────────────────────
    ws_adv = wb.create_sheet("Advertencias")
    _set_header_row(ws_adv, ["#", "Advertencia", "Tipo", "Acción recomendada"], _ORANGE)
    warnings = manifest.get("warnings_activos") or []
    if not warnings:
        warnings = ["Sin advertencias activas en este paquete."]
    for r, w in enumerate(warnings, 2):
        ws_adv.cell(row=r, column=1, value=r - 1)
        ws_adv.cell(row=r, column=2, value=w)
        ws_adv.cell(row=r, column=3, value="Calidad de datos / Validación documental")
        ws_adv.cell(row=r, column=4, value="Verificar fuente antes de presentar")
    _auto_width(ws_adv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 06 Plan de Implementación / Gantt ────────────────────────────────────────

def render_gantt_xlsx(
    manifest: dict,
    resultados: Optional[dict] = None,
    theme_zm: str = "",
    theme_municipio: str = "",
    gantt_plan: Optional[object] = None,
) -> bytes:
    """
    Genera Gantt XLSX con hojas: Fases, Etapas, Hitos, Responsables, Riesgos, Dependencias.
    Si gantt_plan (GanttPlan) está presente, usa fases/etapas reales del Gantt Maestro.
    """
    import openpyxl
    from app.export.gantt_hierarchy import build_hierarchy
    from app.planning.builder import GANTT_PHASES

    wb = openpyxl.Workbook()
    zm = manifest.get("zm") or theme_zm

    # ── Hoja: Fases ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Fases"
    _set_header_row(ws,
        ["#", "Fase ID", "Fase", "Descripción", "Sem inicio", "Sem fin", "Estado"], _GREEN)

    if gantt_plan is not None:
        hierarchy = build_hierarchy(gantt_plan)
        for r, fase in enumerate(hierarchy, 2):
            sem_ini = min((e.inicio_semana for e in fase.etapas), default=1)
            sem_fin = max((e.inicio_semana + e.duracion_semanas for e in fase.etapas), default=1)
            for col, val in enumerate(
                [r - 1, fase.phase_id, fase.nombre, fase.descripcion, sem_ini, sem_fin, "Pendiente"], 1
            ):
                ws.cell(row=r, column=col, value=val)
    else:
        for r, phase_def in enumerate(GANTT_PHASES, 2):
            ws.cell(row=r, column=1, value=r - 1)
            ws.cell(row=r, column=2, value=phase_def["id"])
            ws.cell(row=r, column=3, value=phase_def["nombre"])
            ws.cell(row=r, column=4, value=phase_def["descripcion"])
            ws.cell(row=r, column=5, value="—")
            ws.cell(row=r, column=6, value="—")
            ws.cell(row=r, column=7, value="Pendiente")
    _auto_width(ws)

    # ── Hoja: Etapas (Gantt Maestro) ──────────────────────────────────────────
    ws_etapas = wb.create_sheet("Etapas")
    _set_header_row(ws_etapas,
        ["Task ID", "Etapa", "Fase", "Sem inicio", "Duración sem", "Responsable", "Crítica"], _BLUE)
    if gantt_plan is not None:
        hierarchy = build_hierarchy(gantt_plan)
        row = 2
        for fase in hierarchy:
            for etapa in fase.etapas:
                ws_etapas.cell(row=row, column=1, value=etapa.task_id)
                ws_etapas.cell(row=row, column=2, value=etapa.nombre)
                ws_etapas.cell(row=row, column=3, value=fase.nombre)
                ws_etapas.cell(row=row, column=4, value=etapa.inicio_semana)
                ws_etapas.cell(row=row, column=5, value=etapa.duracion_semanas)
                ws_etapas.cell(row=row, column=6, value=etapa.responsable)
                ws_etapas.cell(row=row, column=7, value="Sí" if etapa.es_critica else "No")
                row += 1
    _auto_width(ws_etapas)

    # ── Hoja: Hitos ───────────────────────────────────────────────────────────
    ws_hitos = wb.create_sheet("Hitos")
    _set_header_row(ws_hitos,
        ["#", "Hito", "Descripción", "Mes", "Gate requerido", "Bloqueante"], _DARK)
    hitos_default = [
        (1, "Reglamento aprobado",           "Aprobación del cabildo con reforma reglamentaria",       3,  "Aprobación cabildo",         "Sí"),
        (2, "Convenio metropolitano firmado", "Acuerdo de coordinación entre municipios de la ZM",     6,  "Firma de presidentes",       "Sí"),
        (3, "Primer CA operativo",           "Centro de acopio en operación con CAS certificado",      18, "Permiso de operación",       "Sí"),
        (4, "Financiamiento confirmado",     "Crédito verde o inversión privada formalizada",          12, "Aprobación tesorería",       "Sí"),
        (5, "50% cobertura municipal",       "Al menos 50% de viviendas en separación de origen",     24, "Meta de cobertura",          "No"),
        (6, "Primer reporte de emisiones",   "CO2e auditado y reportado ante autoridad ambiental",    30, "Certificación CO2e",         "No"),
    ]
    for r, row_data in enumerate(hitos_default, 2):
        for col, val in enumerate(row_data, 1):
            ws_hitos.cell(row=r, column=col, value=val)
    _auto_width(ws_hitos)

    # ── Hoja: Responsables ────────────────────────────────────────────────────
    ws_resp = wb.create_sheet("Responsables")
    _set_header_row(ws_resp,
        ["Rol", "Organismo", "Municipio / ZM", "Función principal", "Nivel de autoridad"], _BLUE)
    responsables_default = [
        ("Coordinador ZM",         "Comisión Metropolitana",            zm,                "Coordinación entre municipios",        "Alta"),
        ("Presidente Municipal",   "Ayuntamiento",                      theme_municipio or zm, "Aprobación normativa y presupuestal", "Alta"),
        ("Director Servicios Púb.","Municipio",                         theme_municipio or zm, "Operación y logística",              "Media"),
        ("Tesorero Municipal",     "Hacienda Municipal",                theme_municipio or zm, "Financiamiento y contratos",         "Alta"),
        ("Concesionario / Operador","Empresa de limpia",                 zm,                "Operación de CAs",                    "Técnica"),
        ("Recicladora ancla",      "Empresa recicladora certificada",   zm,                "Recepción y procesamiento",           "Técnica"),
        ("Contralor Municipal",    "Contraloría",                       theme_municipio or zm, "Auditoría y control",               "Supervisión"),
    ]
    for r, row_data in enumerate(responsables_default, 2):
        for col, val in enumerate(row_data, 1):
            ws_resp.cell(row=r, column=col, value=val)
    _auto_width(ws_resp)

    # ── Hoja: Riesgos ─────────────────────────────────────────────────────────
    ws_riesgos = wb.create_sheet("Riesgos")
    _set_header_row(ws_riesgos,
        ["ID", "Riesgo", "Categoría", "Probabilidad", "Impacto", "Mitigación"], _ORANGE)
    riesgos_default = [
        ("R01", "Cambio de administración municipal",        "Político",       "Alta",  "Alta",   "Convenio y compromisos firmados antes de cambio"),
        ("R02", "Falta de financiamiento",                   "Financiero",     "Media", "Alta",   "Estructurar con crédito verde y participación privada"),
        ("R03", "Baja participación ciudadana",              "Social",         "Media", "Alta",   "Campaña de comunicación y ECAP"),
        ("R04", "Precio de materiales a la baja",            "Mercado",        "Media", "Media",  "Diversificar ingresos con biogás y carbono"),
        ("R05", "Conflicto con pepenadores",                 "Social",         "Alta",  "Media",  "Formalización con ECAP y derechos de trabajo"),
        ("R06", "Retraso en permisos de construcción",       "Regulatorio",    "Media", "Media",  "Gestión anticipada de trámites"),
        ("R07", "Score legal bajo en municipio",             "Jurídico",       "Alta",  "Alta",   "Reforma reglamentaria priorizada antes de cualquier operación"),
        ("R08", "Reglamentos no homologados entre municipios","Jurídico",      "Alta",  "Alta",   "Convenio marco de homologación ZM"),
    ]
    for r, row_data in enumerate(riesgos_default, 2):
        for col, val in enumerate(row_data, 1):
            ws_riesgos.cell(row=r, column=col, value=val)
    _auto_width(ws_riesgos)

    # ── Hoja: Dependencias ────────────────────────────────────────────────────
    ws_dep = wb.create_sheet("Dependencias")
    _set_header_row(ws_dep,
        ["Elemento", "Depende de", "Tipo de dependencia", "Riesgo si falla"], _DARK)
    deps_default = [
        ("Operación de CAs",         "Reforma reglamentaria aprobada",    "Jurídica — bloqueante", "Sin base legal para operar"),
        ("Convenio metropolitano",   "Aprobación de todos los cabildos",  "Política",              "No hay coordinación de tarifas"),
        ("Financiamiento",           "VPN positivo y TIR verificada",     "Financiera",            "Sin proyecto bancable"),
        ("Expansión ZM",             "Piloto municipal exitoso",           "Operativa",             "Sin evidencia para escalar"),
        ("Reporte de emisiones",     "Operación con trazabilidad",        "Ambiental",             "Sin datos auditables"),
    ]
    for r, row_data in enumerate(deps_default, 2):
        for col, val in enumerate(row_data, 1):
            ws_dep.cell(row=r, column=col, value=val)
    _auto_width(ws_dep)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
