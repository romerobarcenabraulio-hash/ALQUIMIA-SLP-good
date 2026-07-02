"""ALQ-18: Export Excel de cobertura municipal — reporte auditable con histórico.

Genera archivos Excel con:
1. Resumen nacional (VERDE/AMARILLO/ROJO counts)
2. Municipios (estado, dimensiones, KPIs, procedencia)
3. Alertas históricas por municipio
4. Análisis de tendencias (cobertura over time)
"""
from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional, List, Dict, Any
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Paleta de Colores (DESIGN_SYSTEM) ──────────────────────────────────────

COLOR_VERDE = "3B6D11"      # Verde verificado
COLOR_AMARILLO = "D4881E"   # Amarillo localizado/estimado
COLOR_ROJO = "C0392B"       # Rojo bloqueado/no disponible
COLOR_GRIS = "A8A49C"       # Gris indefinido
COLOR_HEADER = "1C1B18"     # Header text
COLOR_BG_HEADER = "F4F2ED"  # Header background


class CoberturaExcelBuilder:
    """Constructor de reportes Excel de cobertura municipal"""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def build_summary_sheet(
        self,
        verde_count: int,
        amarillo_count: int,
        rojo_count: int,
        generated_at: Optional[datetime] = None,
    ) -> None:
        """Crea sheet de resumen nacional"""
        ws = self.wb.create_sheet("Resumen Nacional", 0)

        # Estilos
        header_fill = PatternFill(start_color=COLOR_BG_HEADER, end_color=COLOR_BG_HEADER, fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER)
        title_font = Font(name="Calibri", size=14, bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Título
        ws["A1"] = "ALQUIMIA · Semáforo de Cobertura Municipal"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        # Fecha generación
        generated_at = generated_at or datetime.now(timezone.utc)
        ws["A2"] = f"Generado: {generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A2"].font = Font(italic=True, size=9)

        # Resumen
        ws["A4"] = "Estado de Cobertura"
        ws["A4"].font = header_font
        ws["A4"].fill = header_fill

        # Headers
        headers = ["Estado", "Municipios", "Porcentaje", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        total = verde_count + amarillo_count + rojo_count

        # Datos
        data = [
            ("VERDE (Verificado)", verde_count, self._pct(verde_count, total), COLOR_VERDE),
            ("AMARILLO (Localizado/Estimado)", amarillo_count, self._pct(amarillo_count, total), COLOR_AMARILLO),
            ("ROJO (Bloqueado/Sin datos)", rojo_count, self._pct(rojo_count, total), COLOR_ROJO),
            ("TOTAL", total, "100%", COLOR_GRIS),
        ]

        for row_idx, (label, count, pct, color) in enumerate(data, 6):
            # Label
            ws.cell(row=row_idx, column=1).value = label
            ws.cell(row=row_idx, column=1).font = Font(bold=True if row_idx == 9 else False)

            # Count
            ws.cell(row=row_idx, column=2).value = count
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

            # Percentage
            ws.cell(row=row_idx, column=3).value = pct
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

            # Color indicator
            color_cell = ws.cell(row=row_idx, column=4)
            color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 5

    def add_municipios_sheet(
        self,
        municipios: List[Dict[str, Any]],
    ) -> None:
        """Crea sheet con detalle de municipios

        Esperado municipios[].keys:
        - municipio_id, nombre
        - demografia, rsu, legal, contrato, presupuesto, operacion (status)
        - rsu_ton_dia, per_capita (float)
        - data_provenance (dict con sources)
        - siguiente_accion (str)
        - agora_bloqueado (bool)
        """
        ws = self.wb.create_sheet("Municipios", 1)

        # Headers
        headers = [
            "Municipio ID",
            "Nombre",
            "Estado Gral.",
            "Demografía",
            "RSU",
            "Legal",
            "Contrato",
            "Presupuesto",
            "Operación",
            "RSU ton/día",
            "Per Cápita",
            "Bloqueado",
            "Próxima Acción",
        ]

        header_fill = PatternFill(start_color=COLOR_BG_HEADER, end_color=COLOR_BG_HEADER, fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Datos
        for row_idx, mun in enumerate(municipios, 2):
            values = [
                mun.get("municipio_id", ""),
                mun.get("nombre", ""),
                mun.get("estado_general", ""),
                mun.get("demografia", ""),
                mun.get("rsu", ""),
                mun.get("legal", ""),
                mun.get("contrato", ""),
                mun.get("presupuesto", ""),
                mun.get("operacion", ""),
                mun.get("rsu_ton_dia", ""),
                mun.get("per_capita", ""),
                "Sí" if mun.get("agora_bloqueado") else "No",
                mun.get("siguiente_accion", ""),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value
                cell.border = border
                if col in [4, 5, 6, 7, 8, 9]:  # Status columns
                    cell.fill = self._get_status_fill(value)
                if col in [10, 11]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right")

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def add_alerts_sheet(
        self,
        alerts: List[Dict[str, Any]],
    ) -> None:
        """Crea sheet con histórico de alertas

        Esperado alerts[].keys:
        - municipio_id, alert_type, severity
        - title, description
        - created_at (datetime)
        - acknowledged, resolved (bool)
        """
        ws = self.wb.create_sheet("Alertas Históricas", 2)

        headers = [
            "Municipio",
            "Tipo de Alerta",
            "Severidad",
            "Título",
            "Descripción",
            "Fecha",
            "Estado",
        ]

        header_fill = PatternFill(start_color=COLOR_BG_HEADER, end_color=COLOR_BG_HEADER, fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, alert in enumerate(alerts, 2):
            status = "Resuelto" if alert.get("resolved") else ("Leído" if alert.get("acknowledged") else "Nuevo")
            severity_color = self._get_severity_color(alert.get("severity", ""))

            values = [
                alert.get("municipio_id", ""),
                alert.get("alert_type", ""),
                alert.get("severity", ""),
                alert.get("title", ""),
                alert.get("description", ""),
                alert.get("created_at", ""),
                status,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value
                if col == 3:  # Severity column
                    cell.fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def to_bytes(self) -> bytes:
        """Retorna el workbook como bytes"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _pct(count: int, total: int) -> str:
        """Calcula porcentaje"""
        if total == 0:
            return "0%"
        return f"{round((count / total) * 100)}%"

    @staticmethod
    def _get_status_fill(status: str) -> PatternFill:
        """Retorna fill según status"""
        status = str(status).lower()
        if status == "verificado":
            color = COLOR_VERDE
        elif status in ["localizado", "estimado"]:
            color = COLOR_AMARILLO
        elif status in ["no_disponible", "bloqueado"]:
            color = COLOR_ROJO
        else:
            color = COLOR_GRIS
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    @staticmethod
    def _get_severity_color(severity: str) -> str:
        """Retorna color según severidad de alerta"""
        severity = str(severity).lower()
        if severity == "critical":
            return COLOR_ROJO
        elif severity == "high":
            return COLOR_AMARILLO
        else:
            return COLOR_GRIS
