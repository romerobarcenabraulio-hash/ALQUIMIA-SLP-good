"""
Tests Fase 4 — Exportación Profesional DOCX/XLSX/PDF

Criterios del spec (03_FASE_4_EXPORTACION_PROFESIONAL.md):
  - DOCX desde DraftDocument con portada + secciones reales
  - XLSX abre con hojas requeridas
  - XLSX contiene fuentes y advertencias
  - render_report.json contiene assets y bloqueos
  - ZIP profesional contiene manifest + render_report + assets
  - Documento bloqueado NO se exporta como "ok"
  - No se generan .txt
  - Checksum cambia si cambia contenido
  - Jurídico municipal no mezcla municipios
"""
import hashlib
import io
import json
import os
import tempfile
import zipfile
import pytest
from pathlib import Path
from unittest.mock import patch


# ─── Fixtures ─────────────────────────────────────────────────────────────────

MANIFEST_BASICO = {
    "bundle_id":       "test-bundle-f4",
    "zm":              "SLP",
    "municipios":      ["slp"],
    "version":         "0.1-borrador",
    "files":           [
        {"filename": "01_Resumen_Ejecutivo_Municipal.md",  "format": "md"},
        {"filename": "02_Modelo_Tecnico_Financiero.md",    "format": "md"},
    ],
    "fuentes_usadas":  ["INEGI", "Simulador ALQUIMIA", "SEMARNAT"],
    "kpis_incluidos":  ["ton_generadas_diarias", "capex_total", "tir"],
    "warnings_activos": ["Score de confianza bajo (62%)"],
    "score_datos":     62.0,
}

RESULTADOS_BASICOS = {
    "tir":                    18.5,
    "tir_equity":             24.3,
    "vpn":                    12_500_000,
    "capex_total":            45_000_000,
    "ebitda":                 3_200_000,
    "margen_ebitda":          28.5,
    "payback_meses":          38,
    "moic":                   2.4,
    "ingresos_brutos":        87_000_000,
    "empleos_directos":       142,
    "empleos_indirectos":     89,
    "co2e_evitadas_anual":    8_400,
    "co2e_evitadas_horizonte": 84_000,
    "rsu_total_ton_dia":      650,
    "ahorro_salud":           1_200_000,
    "derrama_total":          23_000_000,
    "score_politico":         71,
}

MD_EJECUTIVO = """# 01 Resumen Ejecutivo Municipal

**Audiencia:** Presidencia Municipal, Cabildo
**Decisión que habilita:** Aprobación de inversión en infraestructura de circularidad
**Estado:** `defendible`
**Fuente:** `llm`
**Versión:** `0.1-borrador`

---

## Contexto del programa

San Luis Potosí genera aproximadamente 650 toneladas de RSU por día.
El programa de circularidad propone capturar el 35% de materiales reciclables.

## Modelo financiero

La TIR del proyecto se estima en 18.5% bajo escenario base.

## Recomendación

Se recomienda proceder con la fase de licitación previo acuerdo de cabildo.
"""

MD_JURIDICO_SLP = """# 03 Diagnóstico Jurídico San Luis Potosí

**Audiencia:** Síndico, Asesor jurídico
**Decisión que habilita:** Reforma reglamentaria en municipio de San Luis Potosí
**Estado:** `revision`
**Fuente:** `template`
**Versión:** `0.1-borrador`

---

## Marco reglamentario vigente en San Luis Potosí

El reglamento de limpia del municipio de San Luis Potosí data de 2018.

## Brechas identificadas

Se identifican 4 brechas críticas que requieren reforma.
"""

MD_BLOQUEADO = """# 07 Fuentes y Trazabilidad

**Audiencia:** Contraloría
**Estado:** `bloqueado`
**Fuente:** `template`

> ⚠️ DOCUMENTO BLOQUEADO — Sin ClaimLedger verificado.
"""


# ─── Grupo 1: DocumentRenderer — DOCX real ───────────────────────────────────

class TestDocumentRenderer:

    def test_render_docx_retorna_bytes_no_vacios(self):
        """render_docx() debe retornar bytes válidos de DOCX."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp", date="2026-04-29")
        doc_meta = {
            "titulo": "Resumen Ejecutivo",
            "audiencia": ["Presidencia Municipal"],
            "decision": "Aprobación de inversión",
            "status": "defendible",
            "source": "llm",
            "version": "0.1-borrador",
            "is_fallback": False,
            "is_bloqueado": False,
            "warnings": [],
        }
        result = render_docx(MD_EJECUTIVO, theme, doc_meta, package_id="test-pkg-001")
        assert isinstance(result, bytes)
        assert len(result) > 1000, "DOCX vacío — debe tener contenido real"

    def test_render_docx_es_zip_valido(self):
        """Un archivo DOCX es internamente un ZIP — debe ser abre como ZIP."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="QRO", municipio="queretaro")
        result = render_docx(MD_EJECUTIVO, theme, {"titulo": "Test", "status": "borrador"})
        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            names = zf.namelist()
        assert "word/document.xml" in names, "DOCX no tiene document.xml — no es DOCX real"

    def test_render_docx_contiene_portada_y_titulo(self):
        """El DOCX debe contener el título del documento en el XML."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp")
        doc_meta = {"titulo": "Resumen Ejecutivo Municipal Test", "status": "defendible"}
        result = render_docx(MD_EJECUTIVO, theme, doc_meta)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()
        assert "Resumen Ejecutivo" in doc_xml, "Título no encontrado en DOCX XML"

    def test_render_docx_fallback_marca_borrador(self):
        """Documento is_fallback=True debe incluir marca BORRADOR AUTOMÁTICO."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="MTY", municipio="monterrey")
        doc_meta = {
            "titulo": "Documento fallback",
            "status": "borrador",
            "is_fallback": True,
            "is_bloqueado": False,
        }
        result = render_docx("# Borrador\n\nContenido de template.", theme, doc_meta)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()
        assert "BORRADOR" in doc_xml or "borrador" in doc_xml.lower(), (
            "Documento fallback debe tener marca de BORRADOR visible"
        )

    def test_render_docx_bloqueado_lleva_marca(self):
        """Documento bloqueado debe tener marca BLOQUEADO en el DOCX."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp")
        doc_meta = {
            "titulo": "Doc bloqueado",
            "status": "bloqueado",
            "is_bloqueado": True,
            "is_fallback": False,
        }
        result = render_docx(MD_BLOQUEADO, theme, doc_meta)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()
        assert "BLOQUEADO" in doc_xml, "Documento bloqueado no tiene marca BLOQUEADO"

    def test_render_docx_juridico_contiene_municipio_correcto(self):
        """El DOCX jurídico no debe mezclar municipios — debe contener SLP, no QRO."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="San Luis Potosí")
        doc_meta = {
            "titulo": "Diagnóstico Jurídico San Luis Potosí",
            "status": "revision",
            "municipio": "San Luis Potosí",
        }
        result = render_docx(MD_JURIDICO_SLP, theme, doc_meta)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()

        assert "San Luis Potos" in doc_xml, "Municipio correcto no está en DOCX"
        # Verificar que NO dice Querétaro (mezcla de municipios)
        assert "Quer" not in doc_xml, "DOCX jurídico mezcla municipios — red flag crítico"

    def test_render_docx_con_claim_ledger(self):
        """Si se provee claim_ledger_rows, deben aparecer en el DOCX."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp")
        claims = [
            ("La TIR se estima en 18.5%", "calculado", "el simulador proyecta"),
            ("650 toneladas de RSU/día",  "estimado",  "se estima que"),
        ]
        result = render_docx(MD_EJECUTIVO, theme, {"titulo": "Test"}, claim_ledger_rows=claims)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()
        assert "ClaimLedger" in doc_xml or "Trazabilidad" in doc_xml, (
            "ClaimLedger no aparece en DOCX"
        )

    def test_no_genera_txt(self):
        """render_docx retorna DOCX, nunca .txt disfrazado."""
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp")
        result = render_docx("# Test\n\nContenido.", theme, {"titulo": "Test"})
        # DOCX inicia con bytes de ZIP (PK signature)
        assert result[:2] == b'PK', "El resultado no es un archivo ZIP/DOCX"


# ─── Grupo 2: SpreadsheetRenderer — XLSX ─────────────────────────────────────

class TestSpreadsheetRenderer:

    def test_render_financial_xlsx_retorna_bytes(self):
        """render_financial_xlsx() retorna bytes no vacíos."""
        from app.export.spreadsheet_renderer import render_financial_xlsx

        result = render_financial_xlsx(MANIFEST_BASICO, RESULTADOS_BASICOS, "SLP", "slp", "pkg-f4-001")
        assert isinstance(result, bytes)
        assert len(result) > 500

    def test_financial_xlsx_tiene_hojas_requeridas(self):
        """XLSX financiero debe tener las 7 hojas del spec."""
        from app.export.spreadsheet_renderer import render_financial_xlsx
        import openpyxl

        result = render_financial_xlsx(MANIFEST_BASICO, RESULTADOS_BASICOS, "SLP")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        hojas = wb.sheetnames
        for requerida in ["Resumen", "Inputs", "Resultados", "Flujo_Anual",
                          "Sensibilidades", "Fuentes", "Advertencias"]:
            assert requerida in hojas, f"Hoja '{requerida}' faltante en XLSX financiero"

    def test_financial_xlsx_contiene_fuentes(self):
        """Hoja Fuentes debe tener las fuentes del manifest."""
        from app.export.spreadsheet_renderer import render_financial_xlsx
        import openpyxl

        result = render_financial_xlsx(MANIFEST_BASICO, RESULTADOS_BASICOS)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Fuentes"]
        valores = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "INEGI" in valores, "Fuente INEGI no encontrada en hoja Fuentes"

    def test_financial_xlsx_contiene_advertencias(self):
        """Hoja Advertencias debe reflejar warnings del manifest."""
        from app.export.spreadsheet_renderer import render_financial_xlsx
        import openpyxl

        result = render_financial_xlsx(MANIFEST_BASICO, RESULTADOS_BASICOS)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Advertencias"]
        adv_vals = [str(ws.cell(row=r, column=2).value or "") for r in range(2, ws.max_row + 1)]
        assert any("Score" in v or "confianza" in v.lower() for v in adv_vals), (
            "Advertencia del manifest no encontrada en hoja Advertencias"
        )

    def test_financial_xlsx_sin_resultados_usa_nd(self):
        """Sin resultados, celdas de KPIs muestran 'N/D' — nunca inventar."""
        from app.export.spreadsheet_renderer import render_financial_xlsx
        import openpyxl

        result = render_financial_xlsx(MANIFEST_BASICO, resultados=None)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Resultados"]
        tir_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and "TIR" in str(row[0]):
                tir_row = row
                break
        assert tir_row is not None, "Fila TIR no encontrada en Resultados"
        assert "N/D" in str(tir_row[1]), "Sin resultados, TIR debe ser N/D"

    def test_gantt_xlsx_tiene_hojas_requeridas(self):
        """XLSX Gantt debe tener Fases, Etapas, Hitos, Responsables, Riesgos, Dependencias."""
        from app.export.spreadsheet_renderer import render_gantt_xlsx
        import openpyxl

        result = render_gantt_xlsx(MANIFEST_BASICO, theme_zm="SLP")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        for requerida in ["Fases", "Etapas", "Hitos", "Responsables", "Riesgos", "Dependencias"]:
            assert requerida in wb.sheetnames, f"Hoja '{requerida}' faltante en XLSX Gantt"

    def test_gantt_xlsx_no_esta_vacio(self):
        """Cada hoja del Gantt debe tener al menos 2 filas (header + datos)."""
        from app.export.spreadsheet_renderer import render_gantt_xlsx
        import openpyxl

        result = render_gantt_xlsx(MANIFEST_BASICO, theme_zm="SLP")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        for nombre in ["Fases", "Hitos", "Responsables"]:
            ws = wb[nombre]
            assert ws.max_row >= 2, f"Hoja '{nombre}' tiene solo 1 fila (vacía)"


# ─── Grupo 3: PdfRenderer ────────────────────────────────────────────────────

class TestPdfRenderer:

    def test_render_pdf_ok_o_bloqueado_con_razon(self):
        """
        render_executive_pdf() retorna (bytes, None) OK o (None, razón) bloqueado.
        Nunca tira excepción sin capturar.
        """
        from app.export.pdf_renderer import render_executive_pdf

        pdf_bytes, reason = render_executive_pdf(
            manifest=MANIFEST_BASICO,
            resultados=RESULTADOS_BASICOS,
            theme_zm="SLP",
            package_id="pkg-pdf-test",
        )
        # Cualquier resultado válido es correcto
        if pdf_bytes is None:
            assert reason is not None and len(reason) > 0, (
                "PDF bloqueado debe tener razón explícita"
            )
        else:
            assert len(pdf_bytes) > 0, "PDF retornó bytes vacíos"
            assert reason is None, "PDF exitoso no debe tener razón de bloqueo"

    def test_render_pdf_no_tira_excepcion(self):
        """El renderer de PDF nunca debe propagar excepciones — siempre retorna tupla."""
        from app.export.pdf_renderer import render_executive_pdf

        # Manifest intencionalmente vacío — no debe romper el pipeline
        pdf_bytes, reason = render_executive_pdf(manifest={}, resultados=None)
        # resultado puede ser (bytes, None) o (None, str) — pero nunca excepción
        assert pdf_bytes is not None or reason is not None


# ─── Grupo 4: PackageRenderer — pipeline completo ────────────────────────────

class TestPackageRenderer:

    def _setup_package(self, tmp_dir: str, package_id: str) -> Path:
        """Crea estructura mínima de paquete en disco para tests."""
        pkg_dir   = Path(tmp_dir) / package_id
        files_dir = pkg_dir / "files"
        pkg_dir.mkdir(parents=True)
        files_dir.mkdir()

        # manifest.json
        (pkg_dir / "manifest.json").write_text(
            json.dumps(MANIFEST_BASICO, ensure_ascii=False), encoding="utf-8"
        )
        # Documentos .md
        (files_dir / "01_Resumen_Ejecutivo_Municipal.md").write_text(
            MD_EJECUTIVO, encoding="utf-8"
        )
        (files_dir / "03_Diagnostico_Juridico_Slp.md").write_text(
            MD_JURIDICO_SLP, encoding="utf-8"
        )
        return pkg_dir

    def test_render_package_genera_render_report(self):
        """render_package() debe generar render_report.json en disco."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-render-001"
                self._setup_package(tmp, pkg_id)

                report = pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)
                report_file = Path(tmp) / pkg_id / "rendered" / "render_report.json"
                assert report_file.exists(), "render_report.json no fue creado en disco"
                data = json.loads(report_file.read_text())
                assert "rendered_assets" in data
                assert "blocked_assets"  in data
                assert "qa_status"       in data

    def test_render_package_genera_docx(self):
        """El pipeline debe generar al menos 1 DOCX desde los .md."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-docx-001"
                self._setup_package(tmp, pkg_id)

                report = pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)
                assert report.has_docx(), (
                    f"Pipeline no generó DOCX. Assets: {[a.filename for a in report.rendered_assets]}"
                )

    def test_render_package_genera_xlsx(self):
        """El pipeline debe generar XLSX financiero y Gantt."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-xlsx-001"
                self._setup_package(tmp, pkg_id)

                report = pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)
                assert report.has_xlsx(), (
                    f"Pipeline no generó XLSX. Assets: {[a.filename for a in report.rendered_assets]}"
                )
                xlsx_files = [a for a in report.rendered_assets if a.format == "xlsx"]
                assert len(xlsx_files) == 2, f"Esperados 2 XLSX, generados {len(xlsx_files)}"

    def test_render_package_genera_professional_zip(self):
        """professional_package.zip debe existir con analisis/ e implementacion/."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-zip-001"
                self._setup_package(tmp, pkg_id)

                pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)

                zip_bytes = pkg_mod.get_professional_zip_bytes(pkg_id)
                assert zip_bytes is not None, "professional_package.zip no generado"

                with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                    names = zf.namelist()
                assert "README.txt" in names, "README.txt no en ZIP"
                assert any(n.startswith("analisis/") for n in names), "carpeta analisis/ ausente"
                assert any(n.startswith("implementacion/") for n in names), "carpeta implementacion/ ausente"
                assert "implementacion/00_Maestro/render_report.json" in names

    def test_render_package_incluye_readme_y_actividades(self):
        """ZIP portafolio incluye README raíz y actividades del Gantt."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-portfolio-001"
                self._setup_package(tmp, pkg_id)

                pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)

                zip_bytes = pkg_mod.get_professional_zip_bytes(pkg_id)
                assert zip_bytes is not None
                with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                    names = zf.namelist()
                assert any("actividades/" in n and n.endswith("README.txt") for n in names), (
                    "Actividades con README.txt no encontradas"
                )
                assert "implementacion/00_Maestro/Gantt_ClickUp_Import.csv" in names

    def test_render_package_assets_tienen_checksum(self):
        """Cada RenderedAsset debe tener checksum SHA-256."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-ck-001"
                self._setup_package(tmp, pkg_id)

                report = pkg_mod.render_package(pkg_id, resultados=RESULTADOS_BASICOS)
                for asset in report.rendered_assets:
                    assert asset.checksum, f"Asset {asset.filename} sin checksum"
                    assert len(asset.checksum) == 64, (
                        f"Checksum de {asset.filename} no es SHA-256"
                    )

    def test_checksum_cambia_si_cambia_contenido(self):
        """Dos renders con contenido diferente producen checksums distintos."""
        import hashlib
        from app.export.document_renderer import render_docx
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="SLP", municipio="slp")
        meta  = {"titulo": "Test", "status": "borrador"}

        bytes_a = render_docx("# Versión A\n\nContenido original.", theme, meta)
        bytes_b = render_docx("# Versión B\n\nContenido diferente modificado.", theme, meta)

        ck_a = hashlib.sha256(bytes_a).hexdigest()
        ck_b = hashlib.sha256(bytes_b).hexdigest()
        assert ck_a != ck_b, "Checksums iguales para contenido diferente — red flag"

    def test_docx_bloqueado_registra_en_blocked_y_qa_no_es_ok(self):
        """
        Anti-regresión: paquete con Markdown **Estado:** `bloqueado` NO puede
        producir qa_status='ok' ni n_bloqueados()=0.

        El documento bloqueado debe aparecer en blocked_assets y en rendered
        (con status='bloqueado') de manera que:
          - report.n_bloqueados() > 0
          - report.qa_status != "ok"
        """
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-blocked-antiregression"
                pkg_dir   = Path(tmp) / pkg_id
                files_dir = pkg_dir / "files"
                pkg_dir.mkdir(parents=True)
                files_dir.mkdir()

                # Manifest con UN solo documento
                manifest_bloqueado = dict(MANIFEST_BASICO)
                manifest_bloqueado["files"] = [
                    {"filename": "07_Trazabilidad.md", "format": "md"}
                ]
                (pkg_dir / "manifest.json").write_text(
                    json.dumps(manifest_bloqueado, ensure_ascii=False), encoding="utf-8"
                )
                # Documento con Estado: bloqueado
                (files_dir / "07_Trazabilidad.md").write_text(
                    MD_BLOQUEADO, encoding="utf-8"
                )

                report = pkg_mod.render_package(pkg_id, resultados=None)

                # Verificación central del smoke test
                assert report.n_bloqueados() > 0, (
                    f"n_bloqueados()=0 con documento bloqueado — falla de trazabilidad. "
                    f"blocked_assets={report.blocked_assets}"
                )
                assert report.qa_status != "ok", (
                    f"qa_status='ok' con documento bloqueado — el pipeline no detectó el bloqueo. "
                    f"qa_status={report.qa_status!r}"
                )

                # El documento además debe aparecer en rendered con status="bloqueado"
                bloqueados_en_rendered = [
                    a for a in report.rendered_assets if a.status == "bloqueado"
                ]
                assert len(bloqueados_en_rendered) > 0, (
                    "Documento bloqueado no aparece en rendered_assets con status='bloqueado'"
                )

                # Debe haber warning de ClaimLedger en render_report
                assert any("ClaimLedger" in w or "trazabilidad" in w.lower()
                           for w in report.warnings), (
                    "Render report no contiene warning de ClaimLedger ausente"
                )

    def test_sin_manifest_retorna_report_failed(self):
        """Sin manifest.json, render_package retorna RenderReport con qa_status='failed'."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ALQUIMIA_PACKAGES_DIR": tmp}):
                import importlib
                import app.export.package_renderer as pkg_mod
                importlib.reload(pkg_mod)

                pkg_id = "pkg-f4-nomainfest"
                (Path(tmp) / pkg_id).mkdir()  # sin manifest.json

                report = pkg_mod.render_package(pkg_id)
                assert report.qa_status == "failed"
                assert len(report.errors) > 0


# ─── Grupo 5: Schemas ────────────────────────────────────────────────────────

class TestExportSchemas:

    def test_rendered_asset_serializable(self):
        from app.export.schemas import RenderedAsset

        asset = RenderedAsset(
            package_id="pkg-001",
            source_document_id="01_resumen_ejecutivo_municipal",
            filename="01_Resumen_Ejecutivo_Municipal.docx",
            format="docx",
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            path="/tmp/rendered/01_Resumen_Ejecutivo_Municipal.docx",
            checksum="a" * 64,
            size_bytes=48_000,
            status="ok",
        )
        data = asset.model_dump()
        assert data["format"] == "docx"
        assert data["status"] == "ok"

    def test_render_report_methods(self):
        from app.export.schemas import RenderedAsset, BlockedAsset, RenderReport

        report = RenderReport(
            package_id="pkg-002",
            rendered_assets=[
                RenderedAsset(package_id="pkg-002", filename="a.docx", format="docx",
                              mime_type="application/docx", path="/tmp/a.docx",
                              checksum="b"*64, size_bytes=1000, status="ok"),
                RenderedAsset(package_id="pkg-002", filename="b.xlsx", format="xlsx",
                              mime_type="application/xlsx", path="/tmp/b.xlsx",
                              checksum="c"*64, size_bytes=2000, status="ok"),
            ],
            blocked_assets=[
                BlockedAsset(filename="c.pdf", format="pdf",
                             reason="reportlab no disponible", code="PDF_BLOQUEADO")
            ],
            qa_status="partial",
        )
        assert report.has_docx() is True
        assert report.has_xlsx() is True
        assert report.has_pdf() is False
        assert report.n_ok() == 2
        assert report.n_bloqueados() == 1

    def test_document_theme_defaults(self):
        from app.export.schemas import DocumentTheme

        theme = DocumentTheme(zm="QRO", municipio="queretaro")
        assert theme.brand_name == "ALQUIMIA · paquete documental"
        assert theme.color_primary == "#3B6D11"
        assert theme.typography == "Times New Roman"
